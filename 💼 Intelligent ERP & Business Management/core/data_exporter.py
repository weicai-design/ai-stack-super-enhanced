"""
数据自动导出功能
支持Excel、CSV、PDF等格式
"""

from typing import Dict, List, Optional, Any
from datetime import datetime
import csv
import io

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    CHART_AVAILABLE = True
except ImportError:
    CHART_AVAILABLE = False

class DataExporter:
    """
    数据导出器
    
    功能：
    1. Excel导出（使用openpyxl或pandas）
    2. CSV导出
    3. PDF报表（HTML转PDF）
    4. 定时导出
    """
    
    def __init__(self):
        self.use_openpyxl = OPENPYXL_AVAILABLE
        self.use_pandas = PANDAS_AVAILABLE and not OPENPYXL_AVAILABLE
    
    async def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "Sheet1",
        title: Optional[str] = None
    ) -> bytes:
        """
        导出到Excel
        
        Args:
            data: 数据列表
            filename: 文件名（可选）
            sheet_name: 工作表名称
            title: 报表标题（可选）
            
        Returns:
            Excel文件字节流
        """
        if not data:
            return b""
        
        if self.use_openpyxl:
            return await self._export_with_openpyxl(data, sheet_name, title)
        elif self.use_pandas:
            return await self._export_with_pandas(data, sheet_name, title)
        else:
            # 降级到CSV格式
            return await self.export_to_csv(data, filename)
    
    async def _export_with_openpyxl(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str,
        title: Optional[str]
    ) -> bytes:
        """使用openpyxl导出Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 设置样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 添加标题
        row = 1
        if title:
            ws.merge_cells(f'A1:{get_column_letter(len(data[0].keys()))}1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            row = 2
        
        # 添加表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 添加数据
        for data_row, item in enumerate(data, row + 1):
            for col, header in enumerate(headers, 1):
                value = item.get(header, '')
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    async def _export_with_pandas(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str,
        title: Optional[str]
    ) -> bytes:
        """使用pandas导出Excel"""
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        output.seek(0)
        return output.read()
    
    async def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> bytes:
        """
        导出到CSV
        
        Args:
            data: 数据列表
            filename: 文件名（可选）
            
        Returns:
            CSV文件字节流
        """
        if not data:
            return b""
        
        headers = list(data[0].keys())
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue().encode('utf-8')
    
    async def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str = "数据报表",
        filename: Optional[str] = None
    ) -> bytes:
        """
        导出到PDF（HTML格式，可打印为PDF）
        
        Args:
            data: 数据列表
            title: 报表标题
            filename: 文件名（可选）
            
        Returns:
            HTML文件字节流（浏览器可打印为PDF）
        """
        if not data:
            return b""
        
        # 生成HTML表格
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                }}
                h1 {{
                    text-align: center;
                    color: #333;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #4caf50;
                    color: white;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{
                    background-color: #f2f2f2;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: right;
                    color: #666;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <table>
                <thead>
                    <tr>
        """
        
        # 表头
        headers = list(data[0].keys())
        for header in headers:
            html += f"<th>{header}</th>"
        
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        
        # 数据行
        for item in data:
            html += "<tr>"
            for header in headers:
                value = item.get(header, '')
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += f"""
                </tbody>
            </table>
            <div class="footer">
                生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            </div>
        </body>
        </html>
        """
        
        return html.encode('utf-8')
    
    async def schedule_export(
        self,
        export_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        定时导出
        
        Args:
            export_config: 导出配置
                - data_source: 数据源
                - format: 格式（excel/csv/pdf）
                - schedule: 定时规则（cron格式）
                - destination: 目标路径或邮箱
                
        Returns:
            调度结果
        """
        # TODO: 实现定时导出（使用APScheduler或类似库）
        return {
            "success": True,
            "schedule_id": f"schedule_{datetime.now().timestamp()}",
            "next_run": datetime.now().isoformat(),
            "message": "定时导出功能开发中，当前仅支持手动导出"
        }
    
    async def export_multiple_sheets(
        self,
        sheets_data: Dict[str, List[Dict[str, Any]]],
        filename: Optional[str] = None,
        title: Optional[str] = None
    ) -> bytes:
        """
        导出多工作表Excel
        
        Args:
            sheets_data: 工作表数据字典 {sheet_name: [data]}
            filename: 文件名（可选）
            title: 主标题（可选）
            
        Returns:
            Excel文件字节流
        """
        if not self.use_openpyxl:
            raise ValueError("多工作表导出需要openpyxl库")
        
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        for sheet_name, data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            
            if data:
                headers = list(data[0].keys())
                # 添加表头
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # 添加数据
                for row_idx, item in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))
                
                # 自动调整列宽
                for col in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for row in ws[column_letter]:
                        try:
                            if len(str(row.value)) > max_length:
                                max_length = len(str(row.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    async def export_with_charts(
        self,
        data: List[Dict[str, Any]],
        chart_config: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        导出带图表的Excel（使用openpyxl的图表功能）
        
        Args:
            data: 数据列表
            chart_config: 图表配置
            
        Returns:
            Excel文件字节流
        """
        if not self.use_openpyxl:
            return await self.export_to_excel(data)
        
        if not CHART_AVAILABLE:
            return await self.export_to_excel(data)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "数据报表"
            
            if data:
                headers = list(data[0].keys())
                # 添加表头
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 添加数据
                for row_idx, item in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))
                
                # 添加图表（如果配置了）
                if chart_config:
                    chart_type = chart_config.get('type', 'bar')
                    data_col = chart_config.get('data_column', 1)
                    label_col = chart_config.get('label_column', 0)
                    
                    if chart_type == 'bar':
                        chart = BarChart()
                    elif chart_type == 'line':
                        chart = LineChart()
                    elif chart_type == 'pie':
                        chart = PieChart()
                    else:
                        chart = BarChart()
                    
                    chart.title = chart_config.get('title', '数据图表')
                    chart.x_axis.title = chart_config.get('x_title', '')
                    chart.y_axis.title = chart_config.get('y_title', '')
                    
                    data_range = Reference(ws, min_col=data_col+1, min_row=1, max_row=len(data)+1)
                    labels = Reference(ws, min_col=label_col+1, min_row=2, max_row=len(data)+1)
                    
                    chart.add_data(data_range, titles_from_data=True)
                    chart.set_categories(labels)
                    
                    # 将图表添加到工作表
                    chart_cell = chart_config.get('chart_cell', f'A{len(data)+3}')
                    ws.add_chart(chart, chart_cell)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
        except ImportError:
            # 如果图表功能不可用，返回普通Excel
            return await self.export_to_excel(data)
    
    async def export_with_template(
        self,
        data: List[Dict[str, Any]],
        template_path: Optional[str] = None,
        template_type: str = "standard"
    ) -> bytes:
        """
        使用模板导出
        
        Args:
            data: 数据列表
            template_path: 模板文件路径（可选）
            template_type: 模板类型（standard/financial/production）
            
        Returns:
            Excel文件字节流
        """
        # 根据模板类型应用不同的样式
        if template_type == "financial":
            # 财务模板：添加合计行、格式化金额
            return await self._export_financial_template(data)
        elif template_type == "production":
            # 生产模板：添加进度条、状态标识
            return await self._export_production_template(data)
        else:
            # 标准模板
            return await self.export_to_excel(data, title="数据报表")
    
    async def _export_financial_template(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """财务模板导出"""
        if not self.use_openpyxl:
            return await self.export_to_excel(data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "财务报表"
        
        # 样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        money_format = '#,##0.00'
        
        if data:
            headers = list(data[0].keys())
            # 表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # 数据
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 如果是金额字段，格式化
                    if 'amount' in header.lower() or 'price' in header.lower() or 'total' in header.lower():
                        cell.number_format = money_format
            
            # 添加合计行
            total_row = len(data) + 2
            ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
            # 计算金额列合计
            for col_idx, header in enumerate(headers, 1):
                if 'amount' in header.lower() or 'price' in header.lower() or 'total' in header.lower():
                    col_letter = get_column_letter(col_idx)
                    ws.cell(row=total_row, column=col_idx, 
                           value=f"=SUM({col_letter}2:{col_letter}{len(data)+1})")
                    ws.cell(row=total_row, column=col_idx).number_format = money_format
                    ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    async def _export_production_template(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """生产模板导出"""
        if not self.use_openpyxl:
            return await self.export_to_excel(data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "生产报表"
        
        # 样式
        header_fill = PatternFill(start_color="4caf50", end_color="4caf50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        status_colors = {
            'completed': PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid"),
            'in_progress': PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid"),
            'pending': PatternFill(start_color="ffccbc", end_color="ffccbc", fill_type="solid")
        }
        
        if data:
            headers = list(data[0].keys())
            # 表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # 数据
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 状态列着色
                    if header.lower() == 'status' and value in status_colors:
                        cell.fill = status_colors[value]
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

